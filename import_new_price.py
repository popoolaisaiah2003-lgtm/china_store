import openpyxl
import re
from app import create_app
from extensions import db
from models import Category, Product, ProductImage

def slugify(text):
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    return text

def parse_price(val):
    if not val:
        return 0.0
    # Clean string, extract numeric price
    s = str(val).strip().replace('$', '').replace('￥', '').replace(',', '')
    # Handle notes in price string like '300（高纯330）'
    match = re.search(r'^\d+(\.\d+)?', s)
    if match:
        return float(match.group(0))
    try:
        return float(s)
    except ValueError:
        return 0.0

def format_spec(spec_str):
    if not spec_str:
        return ""
    # Convert '5mg*10vials' -> '5mg (10 vials)'
    # Convert '10mg *10 vials' -> '10mg (10 vials)'
    s = spec_str.strip()
    match = re.match(r'^([\d\.\w]+)\s*[*xX]\s*([\d\.\w\s]+)$', s)
    if match:
        dosage = match.group(1).strip()
        pack = match.group(2).strip()
        if 'vial' in pack.lower() and not pack.lower().endswith('s'):
            pack = pack + 's'
        if not ('vial' in pack.lower() or 'bottle' in pack.lower()):
            pack = pack + ' vials'
        return f"{dosage} ({pack})"
    return s

def categorize(name, cat_no=""):
    n = name.upper()
    
    # 1. GLP-1 & Metabolic
    if any(k in n for k in ['SEMAGLUTIDE', 'TIRZEPATIDE', 'RETATRUTIDE', 'CAGRILINTIDE', 'MAZDUTIDE', 'SURVODUTIDE', 'GLP-1', 'AOD9604', '5-AMINO-1MQ', 'ADIPOTIDE', 'AICAR', 'LIPO', 'CARNITINE', 'SLU-PP-332']):
        return 'Metabolic & GLP-1 Analogues'
        
    # 2. Tissue Recovery & Healing
    if any(k in n for k in ['BPC', 'TB500', 'THYMOSIN B4', 'KPV', 'ARA-290', 'THYMOSIN ALPHA', 'THYMALIN', 'GLOW', 'KLOW']):
        return 'Tissue Recovery & Healing'
        
    # 3. GH Secretagogues
    if any(k in n for k in ['HGH', 'SOMATROPIN', 'CJC', 'IPAMORELIN', 'LPAMORELIN', 'GHRP', 'HEXARELIN', 'SERMORELIN', 'TESAMORELIN', 'FRAGMENT']):
        return 'GH Secretagogues'
        
    # 4. Pigmentation & Wellness
    if any(k in n for k in ['MT-1', 'MT-2', 'MELANOTAN', 'MELANOTIN', 'PT-141', 'GHK', 'NAD', 'GLUTATHIONE', 'EPITHALON', 'MOTS', 'SELANK', 'SEMAX', 'OXYTOCIN', 'DSIP', 'MELATONIN']):
        return 'Pigmentation & Wellness'
        
    # 5. Other Peptides
    return 'Other Peptides'

def run_import():
    excel_path = r"C:\Users\HomePC\Desktop\ads\new_price.xlsx"
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active

    app = create_app()
    with app.app_context():
        print("[Import] Recreating DB schema...")
        db.drop_all()
        db.create_all()

        # Create 5 Target Categories
        cat_names = [
            'Metabolic & GLP-1 Analogues',
            'Tissue Recovery & Healing',
            'GH Secretagogues',
            'Pigmentation & Wellness',
            'Other Peptides'
        ]
        
        cat_map = {}
        for cname in cat_names:
            c = Category(name=cname, slug=slugify(cname), description=f"Yan Zhen Wholesale {cname}.")
            db.session.add(c)
            db.session.flush()
            cat_map[cname] = c.id

        db.session.commit()

        imported_count = 0
        current_base_name = ""

        # Skip header (row 1)
        rows = list(sheet.iter_rows(values_only=True))[1:]

        for idx, row in enumerate(rows, start=2):
            if not row or not any(row):
                continue

            code = str(row[0]).strip() if row[0] is not None else ""
            raw_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            raw_spec = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            raw_price = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""

            # Carry forward base product name if blank in current row
            if raw_name:
                current_base_name = raw_name

            base_name = current_base_name or code or f"Peptide-{idx}"
            formatted_spec = format_spec(raw_spec)

            # Construct full product title according to format requirement
            # e.g., Semaglutide 5mg (10 vials)
            if formatted_spec:
                full_product_name = f"{base_name} {formatted_spec}"
            else:
                full_product_name = base_name

            price = parse_price(raw_price)

            slug = slugify(f"{code}-{full_product_name}")
            if not slug or len(slug) < 2:
                slug = slugify(f"prod-{idx}-{full_product_name}")

            # Ensure unique slug
            existing = Product.query.filter_by(slug=slug).first()
            if existing:
                slug = f"{slug}-{idx}"

            category_name = categorize(base_name, code)
            category_id = cat_map[category_name]

            product = Product(
                category_id=category_id,
                name=full_product_name,
                slug=slug,
                purity='>= 99.8%',
                sequence_or_cas=code if code else None,
                short_description=f"Specification: {formatted_spec}" if formatted_spec else "Wholesale research peptide vial.",
                description=f"Authentic Yan Zhen Wholesale {full_product_name}. Analytical HPLC batch-certified research peptide vial.",
                price=price,
                stock_status='In Stock',
                is_featured=(imported_count < 8)
            )
            db.session.add(product)
            imported_count += 1

        db.session.commit()
        print(f"[Import Success] Successfully imported {imported_count} products from new_price.xlsx!")

if __name__ == '__main__':
    run_import()
