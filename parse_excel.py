import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\HomePC\Desktop\ads\new_price.xlsx"
wb = openpyxl.load_workbook(excel_path)
sheet = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active

for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
    vals = [str(c).strip() if c is not None else "" for c in row[:5]]
    if any(vals) and row_idx <= 64:
        print(f"Row {row_idx}: ColA='{vals[0]}', ColB='{vals[1]}', ColC='{vals[2]}', ColD='{vals[3]}'")
